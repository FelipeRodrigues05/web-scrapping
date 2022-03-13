from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By


class WebScrapping:

    product_name_list  = []
    product_price_list = []
    # configs do Chrome
    def __init__(self):
        chrome_options = Options()
        chrome_options.binary_location = "/usr/bin/google-chrome"
        self.driver = webdriver.Chrome("/usr/local/bin/chromedriver")
        print("Configuração Completa!")

    # Chama as funções
    def Init(self):
        self.next_page = 1
        self.access_site()

    def access_site(self):
        self.driver.get("https://telefonesimportados.netlify.app")
        self.capture_informations()

    def capture_informations(self):
        for p in range(5):
            item = 1
            for i in range(12):
                products = self.driver.find_elements_by_xpath(
                        f'/html/body/div[5]/div[2]/div[1]/div[{item}]/div/h2/a')
                self.product_name_list.append(products[0].text)

                prices = self.driver.find_elements_by_xpath(
                        f'//div[{item}]/div[@class="single-shop-product" and 1]/div[@class="product-carousel-price" and 2]/ins[1]')
                self.product_price_list.append(prices[0].text)
                item += 1

            self.save_information()

            try:
                next_button = self.driver.find_element_by_xpath(
                        '/html/body/div[5]/div[2]/div[2]/div/div/nav/ul/li[7]/a')
                next_button.click()
                print("Navegando para a próxima página")

            except:
                print("Sem mais páginas")
                self.driver.quit()

    def save_information(self):
        index = 2
        sheet = Workbook()
        cellphones = sheet["Sheet"]
        cellphones.title = "Cellphones"
        cellphones['A1'] = "Name"
        cellphones['B1'] = "Price"

        for name, price in zip(self.product_name_list, self.product_price_list):
            cellphones.cell(column=1, row=index, value=name)
            cellphones.cell(column=2, row=index, value=price)
            index += 1
        print("Planilha Criada com sucesso")
        sheet.save("products.xlsx")
